import os
import pandas as pd
import openpyxl
import model.dataholder as dh


print(os.getcwd())
os.chdir('./data')


fileName = 'WochenübersichtMV_2022.xlsx'
data = pd.ExcelFile(fileName)
print(data.sheet_names) #this returns the all the sheets in the excel file

workbook = openpyxl.load_workbook(fileName)
worksheet1 = workbook.worksheets[0]

holder = dh.DataHolder("savefile.json")
for row in range(1, worksheet1.max_row + 1):
    holder.data[row] = {}
    for col in range(1, worksheet1.max_column + 1):
        holder.data[row][col] = worksheet1.cell(row, col).value

holder.save()

br = 0
